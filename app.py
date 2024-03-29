import gmail
import requests
# import ConfigParser
import configparser
import os
import html2text

from library.config_parser import configSectionParser
from openpyxl import load_workbook, worksheet
from flask import Flask
from flask import jsonify

#INIT CONFIG
config = configparser.ConfigParser()
config.read("config/config.cfg")

#CONFIG
receiver_email = configSectionParser(config,"RECEIVER")['email']
receiver_password = configSectionParser(config,"RECEIVER")['password']
sender_email = configSectionParser(config,"SENDER")['email']
emailer_url = configSectionParser(config,"EMAILER")['url']

# # INIT
# g = gmail.login(receiver_email, receiver_password)
# # filter email based on sender
# em = g.inbox()

# emails = g.inbox().mail(sender=sender_email, prefetch=True, unread=True)

application = app = Flask(__name__)

# INDEX 
@app.route('/')
def index():
    g = gmail.login(receiver_email, receiver_password)

    if not g:
        print("Login Failed")
    else:
        unread = g.inbox().mail(sender=sender_email, prefetch=True, unread=True)
        print(len(unread))
        for message in unread:
            if message.subject == "Invitation":
                print("Message", message.subject)
                print(locate_password(message.body))
                password = locate_password(message.body)

            # if message.subject == "Re-Invitation":
            #     password = locate_password(message.body)    
            #     print("Message", message.subject)
            #     print(locate_password(message.body))

        return password
    return "nani"

def locate_password(email_content):
    assert email_content, "Email Content is required."

    format_message = html2text.html2text(str(email_content, 'UTF8'))
    list_message = format_message.split()

    # Return next word or password
    password = list_message[list_message.index('password:') + 1]

    return password
# CHECKER
@app.route('/email/checker')
def email_checker():
    for email in emails:
        data =  email.body
        for attachment in email.attachments:
            attachment.save(attachment.name)
            return file_reader(attachment.name)

def file_reader(file_name):

    wb = load_workbook(filename = file_name)
    ws = wb.active

    array = []
    final_array = []
    row_length = 0

    # GET ROW LENGTH
    for row in ws.iter_rows():
        row_length = len(row)
        break

    # GET VALUES PER ROW
    for row in ws.iter_rows():
        cel_values = []
        for cell in row:
            try:
                cel_values.append(str(cell.value))
            except:
                cel_values.append(cell.value)

        array.append(cel_values)

    _dict = {}

    arr = array[0]

    # REMOVE FIRST INDEX
    array.pop(0)

    # LOOP DATA AND APPEND TO DICTIONARY
    for array in array:
        ctr = 0
        for key in arr:
            key = key.lower()
            if ctr > row_length: break

            _dict[key] = array[ctr]

            ctr += 1

        final_array.append(_dict.copy())
    
    # LOOP DATA
    for remittance in final_array:
        payload = { 
            "serviceType": "Remittance", 
            "fspName": "Sendah", 
            "email": remittance['email'], 
            "subject": "Remittance", 
            "name": remittance['sender name'], 
            "type": "success-remittance", 
            "amount":remittance['amount'], 
        }
        # CALL SEND EMAIL
        r = requests.post(emailer_url, json=payload)
        print(r.text)


    try:
        os.remove(file_name)
    except OSError:
        pass
    return jsonify(result=final_array) 

if __name__ == '__main__':
    #app.run()
    app.run(host='0.0.0.0', port=5000)
